#!/bin python
# -*- coding: utf-8 -*-
# @Time    : 2020/10/23 15:57
# @Author  : ZhangL

import openpyxl
import xlrd

insure_workbook = xlrd.open_workbook("计划书.xlsx")
insure_sheet = insure_workbook.sheet_by_index(0)

insure_list = []
for i in range(1, insure_sheet.nrows):
    insure_list.append(insure_sheet.cell_value(i, 0))

product_workbook = openpyxl.load_workbook('产品S.xlsx')
product_sheet = product_workbook['清单']

for i in range(2, product_sheet.max_row):
    product_name = product_sheet.cell(i, 4).value
    product_name_list = []
    for insure_name in insure_list:
        if insure_name in product_name:
            product_name_list.append(insure_name)
    if len(product_name_list) > 0:
        product_sheet.cell(i, 12).value = ','.join(product_name_list)

product_workbook.save("result.xlsx")
