# import unittest
from openpyxl import load_workbook


class ParseExcel(object):
    def __init__(self, excelPath):
        # 将要读取的excel加载到内存上
        self.wb = load_workbook(excelPath)
        # 通过工作表名称获取一个工作表对象
        self.sheet = self.wb.worksheets[0]
        # 获取工作表中存在数据的区域的最大行号
        self.maxRowNum = self.sheet.max_row
        self.rowNum=self.sheet.max_row
        self.colNum=self.sheet.max_column
        # print("这是行：",self.rowNum)
        # print("这是列：",self.colNum)

    def getDatasFromSheet(self):
        # 用于存放从工作表中读取的数据
        dataList = []
        for row in self.sheet.rows:
            # 遍历工作表中数据区域的每一行
            # 并将每行中各个单元格的数据取出存于列表tmpList中，
            # 然后再讲存放一行数据的列表添加到最终数据列表dataList中
            # print(row)
            tmpList = []
            for cell in row:
                # print(type(cell.value))
                tmpList.append(cell.value)
                # print(cell.value)
            # print("下面是一行单元格的值")
            # print(tmpList)
            dataList.append(tmpList)
            #print(dataList)
            # 将获取工作表中的所有数据的迭代对象返回
        dataList.pop(0)
        return dataList


if __name__ == '__main__':
    excelPath = "../data/汇缴补缴明细.xlsx.xlsx"
    pe = ParseExcel(excelPath)
    result = pe.getDatasFromSheet()
    print(result)

    # for i in result:
    #     print(i)
    #     print(i[0], i[1][2])
   #  unittest.main()
